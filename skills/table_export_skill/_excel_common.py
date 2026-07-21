"""
Excel 导出公共工具模块
提取三个导出文件共享的配置加载、表头样式、格式化逻辑
"""

import os
import sys
from typing import List, Optional

_project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)


def load_template_config(template_path: Optional[str] = None) -> dict:
    """读取 table_template.yaml 模板配置"""
    if template_path is None:
        template_path = os.path.join(_project_root, 'config', 'table_template.yaml')
    try:
        import yaml
        with open(template_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception:
        raise FileNotFoundError(f"无法读取模板配置文件: {template_path}")


def load_table_format() -> dict:
    """读取全局配置中的 table_format 设置"""
    try:
        import yaml
        global_config_path = os.path.join(_project_root, 'config', 'global_config.yaml')
        with open(global_config_path, 'r', encoding='utf-8') as f:
            global_config = yaml.safe_load(f)
        return global_config.get("table_format", {})
    except Exception:
        return {}


def setup_header_style(table_format: dict):
    """生成统一的表头字体和填充样式，返回 (font, fill) 元组"""
    from openpyxl.styles import Font, PatternFill
    bold = table_format.get("header_font_bold", True)
    font = Font(bold=bold, color="FFFFFF", size=11)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    return font, fill


def write_headers(ws, columns: list, font, fill):
    """写入表头行，提取 headers 和 fields 列表"""
    from openpyxl.styles import Alignment
    headers = [col.get("header", "") for col in columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return headers


def apply_worksheet_formatting(ws, columns: list, data_row_count: int, table_format: dict):
    """统一处理冻结首行、自动筛选、自动列宽"""
    from openpyxl.utils import get_column_letter

    # 冻结首行
    if table_format.get("freeze_first_row", True):
        ws.freeze_panes = "A2"

    # 自动筛选
    if table_format.get("add_auto_filter", True):
        max_col = len(columns)
        max_row = data_row_count + 1
        if max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 自动列宽
    if table_format.get("auto_column_width", True):
        max_column_width = table_format.get("max_column_width", 60)
        for col_idx, col_config in enumerate(columns, 1):
            width = min(col_config.get("width", 20), max_column_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width