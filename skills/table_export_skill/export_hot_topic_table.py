"""
热点话题 Excel 表格导出模块
将话题聚类结果转换为格式化的 Excel 文件
"""

import os
from typing import Dict, List, Optional

from ._excel_common import (
    load_template_config,
    load_table_format,
    setup_header_style,
    write_headers,
    apply_worksheet_formatting,
)


def export_hot_topic_table(
    topics: Dict[int, dict],
    output_path: str,
    template_path: Optional[str] = None,
    video_id: str = "",
) -> str:
    """
    将话题聚类结果展平为 Excel 表格

    Args:
        topics: cluster_topics 返回的话题聚类结果字典
        output_path: 输出 Excel 文件路径
        template_path: 模板配置文件路径，默认读取 config/table_template.yaml
        video_id: 关联的视频 ID

    Returns:
        输出文件的绝对路径
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        raise ImportError("需要安装 openpyxl (pip install openpyxl)")

    template = load_template_config(template_path)
    table_format = load_table_format()

    topic_config = template.get("comment_hot_topic", {})
    columns = topic_config.get("columns", [])
    fields = [col.get("field", "") for col in columns]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = topic_config.get("sheet_name", "评论热点问题")

    font, fill = setup_header_style(table_format)
    write_headers(ws, columns, font, fill)

    data_row_count = 0
    if not topics:
        empty_row = {field: "" for field in fields}
        empty_row["video_id"] = video_id
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=2, column=col_idx, value=empty_row.get(field, ""))
        data_row_count = 1
    else:
        total_comment_num = sum(
            len(t.get("comment_ids", [])) for t in topics.values()
        )
        row_idx = 2
        for topic_id, topic_data in topics.items():
            keyword_str = "、".join(topic_data.get("keywords", []))
            comment_ids = topic_data.get("comment_ids", [])
            total_num = len(comment_ids)
            topic_weight = total_num / total_comment_num if total_comment_num > 0 else 0
            sample_question = f"关联 {total_num} 条评论" if comment_ids else ""

            row_data = {
                "video_id": video_id,
                "topic_id": topic_id,
                "topic_keyword": keyword_str,
                "sample_question": sample_question,
                "total_comment_num": total_num,
                "avg_like": 0.0,
                "topic_weight": round(topic_weight, 4),
                "related_comment_ids": ",".join(comment_ids[:10]),
            }
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
                cell.alignment = Alignment(vertical="center")
            row_idx += 1
            data_row_count += 1

    apply_worksheet_formatting(ws, columns, data_row_count, table_format)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)

    return os.path.abspath(output_path)